# Excel
from openpyxl import load_workbook

# Google API
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Utils
import os.path


SCOPES = ['https://www.googleapis.com/auth/calendar']

def create_events():
    """Create events for Google Calendar API
    Create event for every friend birthday in list of friends
    """
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    list_friends, emails = read_excel()
    for friend in list_friends:
        event = {
            'summary': 'Cumpleaños de %s' % (friend['name']),
            'location': '%s, %s' % (friend['city'], friend['state']),
            'description': 'Recordatorio para que le saludes por su cumple a %s' % (friend['name']),
            'start': {
                'dateTime': '2022-%s-%sT09:00:00-05:00' % (friend['month'], friend['day']),
                'timeZone': 'America/Lima',
            },
            'end': {
                'dateTime': '2022-%s-%sT23:00:00-05:00' % (friend['month'], friend['day']),
                'timeZone': 'America/Lima',
            },
            'recurrence': [
                'RRULE:FREQ=YEARLY'
            ],
            'attendees': emails,
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'email', 'minutes': 1 * 60},
                    {'method': 'popup', 'minutes': 10},
                ],
            },
        }
        try:
            service = build('calendar', 'v3', credentials=creds)
            event = service.events().insert(calendarId='primary', body=event).execute()
            print ('Event created: %s' % (event.get('htmlLink')))

        except HttpError as error:
            print('An error occurred: %s' % error)


def read_excel():
    wb = load_workbook(filename = 'DrinkTim.xlsx')
    ws = wb['Hoja1']
    list_friends = []
    emails = []
    for row in ws.iter_rows(min_row=2):
        new_dict = {
            'name': row[4].value,
            'day': str(row[3].value) if len(str(row[3].value)) == 2 else '0' + str(row[3].value),
            'month': str(row[2].value) if len(str(row[2].value)) == 2 else '0' + str(row[2].value),
            'city': row[6].value,
            'state': row[7].value
        }
        list_friends.append(new_dict)
        if row[5].value:
            emails.append({'email': row[5].value})
    return list_friends, emails

if __name__ == '__main__':
    create_events()